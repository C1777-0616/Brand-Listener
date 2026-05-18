import requests
import gzip
import json
from typing import Optional, Dict, Any, List
from datetime import datetime


class XHSNoteDetail:
    """小红书笔记详情采集类"""
    
    def __init__(self, api_token: str):
        """
        初始化笔记详情采集类
        
        Args:
            api_token: API 访问令牌
        """
        self.base_url = "https://proxy-api.ainm.store"
        self.api_token = api_token
        # 使用 get-note-detail/v1 API 获取笔记详情
        self.endpoint = "/p/xiaohongshu/get-note-detail/v1"
    
    def get_note_detail(
        self,
        note_id: str
    ) -> Dict[str, Any]:
        """
        获取单个笔记的详细信息
        
        Args:
            note_id: 笔记 ID
        
        Returns:
            包含笔记详情的字典
        """
        return self._fetch_note_detail(note_id)
    
    def _fetch_note_detail(
        self,
        note_id: str
    ) -> Dict[str, Any]:
        """
        获取笔记详情
        
        Args:
            note_id: 笔记 ID
        
        Returns:
            API 响应结果
        """
        url = f"{self.base_url}{self.endpoint}"
        
        headers = {
            "Authorization": f"Bearer {self.api_token}"
        }
        
        params = {
            "noteId": note_id
        }
        
        try:
            response = requests.get(url, params=params, headers=headers, timeout=60)
            response.raise_for_status()
            
            # 解析响应（支持 gzip 压缩）
            result = self._parse_response(response)
            
            # 修复：API 返回 code: 200 表示成功，不是 0
            if result and result.get("code") in [0, 200]:
                # 修复：处理嵌套的 data 字段
                outer_data = result.get("data", {})
                
                # 如果有嵌套的 data 字段，使用内层 data
                if isinstance(outer_data, dict) and "data" in outer_data:
                    inner_data = outer_data.get("data", [])
                else:
                    inner_data = outer_data
                
                # 返回数据在 data 数组中
                if isinstance(inner_data, list) and len(inner_data) > 0:
                    note_data = inner_data[0]
                    
                    # 修复：笔记数据在 note_list 数组中
                    note_list = note_data.get("note_list", [])
                    if note_list and isinstance(note_list, list) and len(note_list) > 0:
                        actual_note = note_list[0]
                    else:
                        # 降级使用 note_data 本身
                        actual_note = note_data
                    
                    return {
                        "success": True,
                        "data": {
                            "user": note_data.get("user", {}),
                            "note": actual_note
                        },
                        "message": "获取成功"
                    }
                else:
                    return {
                        "success": False,
                        "error_code": "NO_DATA",
                        "message": "未找到笔记数据"
                    }
            else:
                return {
                    "success": False,
                    "error_code": result.get("code"),
                    "message": self._get_error_message(result.get("code"))
                }
        
        except requests.exceptions.Timeout:
            return {
                "success": False,
                "error_code": "TIMEOUT",
                "message": "请求超时，建议重试"
            }
        except requests.exceptions.RequestException as e:
            return {
                "success": False,
                "error_code": "NETWORK_ERROR",
                "message": f"网络错误：{str(e)}"
            }
        except json.JSONDecodeError:
            return {
                "success": False,
                "error_code": "PARSE_ERROR",
                "message": "响应解析失败"
            }
    
    def get_notes_by_ids(
        self,
        note_ids: List[str],
        delay_seconds: float = 1.0
    ) -> Dict[str, Any]:
        """
        批量获取多个笔记的详情
        
        Args:
            note_ids: 笔记 ID 列表
            delay_seconds: 请求延迟时间
        
        Returns:
            包含所有笔记详情的字典
        """
        import time
        
        all_notes = []
        
        for i, note_id in enumerate(note_ids, 1):
            result = self._fetch_note_detail(note_id)
            
            if result.get("success"):
                note = result.get("data", {}).get("note", {})
                if note:
                    all_notes.append(note)
            else:
                print(f"❌ 获取笔记 {note_id} 失败：{result.get('message')}")
            
            # 延迟以避免限流
            if i < len(note_ids):
                time.sleep(delay_seconds)
        
        return {
            "success": True,
            "data": {
                "notes": all_notes,
                "total_found": len(all_notes),
                "total_requested": len(note_ids)
            },
            "message": f"成功获取 {len(all_notes)}/{len(note_ids)} 条笔记"
        }
    
    def _get_error_message(self, code: int) -> str:
        """获取错误码对应的中文说明"""
        error_messages = {
            100: "Token 无效或已失效",
            301: "采集失败，请重试",
            302: "超出速率限制",
            303: "超出每日配额",
            400: "参数错误",
            500: "内部服务器错误",
            600: "权限不足",
            601: "余额不足"
        }
        return error_messages.get(code, f"未知错误 (code: {code})")
    
    @staticmethod
    def _parse_response(response) -> Optional[Dict[str, Any]]:
        """解析响应，支持 gzip 压缩"""
        raw = response.content
        
        try:
            # 尝试 gzip 解压
            decompressed = gzip.decompress(raw)
            return json.loads(decompressed.decode('utf-8'))
        except:
            # 尝试直接解码
            try:
                return json.loads(raw.decode('utf-8'))
            except:
                return None
    
    def format_note_detail(self, note_result: Dict[str, Any]) -> str:
        """格式化输出笔记详情"""
        if not note_result.get("success"):
            return f"❌ 获取失败：{note_result.get('message')}"
        
        data = note_result.get("data", {})
        note = data.get("note", {})
        
        if not note:
            return "ℹ️ 未找到笔记"
        
        user = note.get("user", {})
        images = note.get("images_list", [])
        
        output = []
        output.append("=" * 80)
        output.append("📝 笔记详情")
        output.append("=" * 80)
        
        note_id = note.get('id', 'N/A')
        output.append(f"\n🆔 笔记 ID: {note_id}")
        
        # 提取 xsec_token 并显示完整链接
        xsec_token = None
        mini_program_info = note.get("mini_program_info", {})
        if not mini_program_info:
            mini_program_info = note.get("qq_mini_program_info", {})
        
        if mini_program_info:
            path = mini_program_info.get("path", "")
            if path:
                import urllib.parse
                # 先 URL 解码
                decoded_path = urllib.parse.unquote(path)
                # 再查找 xsec_token
                if "xsec_token=" in decoded_path:
                    xsec_token = decoded_path.split("xsec_token=")[1].split("&")[0]
        
        if xsec_token:
            note_url = f"https://www.xiaohongshu.com/explore/{note_id}?xsec_token={xsec_token}"
            output.append(f"🔗 笔记链接：{note_url}")
        else:
            output.append(f"🔗 笔记链接：https://www.xiaohongshu.com/explore/{note_id} (缺少 xsec_token)")
        output.append(f"📌 标题：{note.get('title', note.get('display_title', 'N/A'))}")
        output.append(f"📄 描述：{note.get('desc', 'N/A')}")
        output.append(f"👤 作者：{user.get('nickname', 'N/A')} (ID: {user.get('userid', 'N/A')})")
        
        output.append(f"\n📊 互动数据:")
        output.append(f"   ❤️  点赞：{note.get('likes', 0)}")
        output.append(f"   ⭐  收藏：{note.get('collected_count', 0)}")
        output.append(f"   💬  评论：{note.get('comments_count', 0)}")
        output.append(f"   🔗  分享：{note.get('share_count', 0)}")
        
        output.append(f"\n📷 媒体信息:")
        output.append(f"   类型：{note.get('type', 'N/A')}")
        output.append(f"   图片数量：{len(images)}")
        
        if images:
            for i, img in enumerate(images[:3], 1):  # 只显示前 3 张
                url = img.get("url", "")[:60]
                output.append(f"   图片{i}: {url}...")
            if len(images) > 3:
                output.append(f"   ... 还有 {len(images) - 3} 张图片")
        
        create_time = note.get("create_time")
        if create_time:
            dt = datetime.fromtimestamp(create_time)
            output.append(f"\n📅 发布时间：{dt.strftime('%Y-%m-%d %H:%M:%S')}")
        
        time_desc = note.get("time_desc")
        if time_desc:
            output.append(f"🕐 时间描述：{time_desc}")
        
        # 标签信息
        desc = note.get("desc", "")
        if "#" in desc:
            import re
            tags = re.findall(r'#([^#]+)#', desc)
            if tags:
                output.append(f"\n🏷️  标签：{', '.join(tags)}")
        
        output.append("=" * 80)
        
        return "\n".join(output)
    
    def export_to_json(self, note_result: Dict[str, Any], filename: str) -> bool:
        """
        导出笔记详情到 JSON 文件
        
        Args:
            note_result: 笔记数据结果
            filename: 输出文件名
        
        Returns:
            是否导出成功
        """
        if not note_result.get("success"):
            return False
        
        try:
            data = note_result.get("data", {})
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return True
        except Exception:
            return False
    
    def export_to_excel(self, notes: list, filename: str) -> bool:
        """
        导出笔记数据到 Excel 文件
        
        Args:
            notes: 笔记数据列表
            filename: 输出文件名
        
        Returns:
            是否导出成功
        """
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment
            
            if not notes:
                print("❌ 没有数据可导出")
                return False
            
            # 提取字段
            data_rows = []
            for note in notes:
                user = note.get('user', {})
                images = note.get('images_list', [])
                
                # 获取第一张图片的 URL（实际链接，非超链接）
                cover_image_url = images[0].get('url', '') if images else ''
                
                # 构建笔记链接
                note_id = note.get('id', '')
                
                # 优先从 mini_program_info 中提取 xsec_token
                xsec_token = None
                mini_program_info = note.get('mini_program_info', {})
                if not mini_program_info:
                    # 尝试从 qq_mini_program_info 获取
                    mini_program_info = note.get('qq_mini_program_info', {})
                
                if mini_program_info:
                    path = mini_program_info.get('path', '')
                    if path:
                        import urllib.parse
                        # 先 URL 解码
                        decoded_path = urllib.parse.unquote(path)
                        # 再查找 xsec_token
                        if 'xsec_token=' in decoded_path:
                            # 提取 xsec_token
                            xsec_token = decoded_path.split('xsec_token=')[1].split('&')[0]
                
                # 构建完整链接
                if xsec_token:
                    note_url = f"https://www.xiaohongshu.com/explore/{note_id}?xsec_token={xsec_token}"
                else:
                    note_url = f"https://www.xiaohongshu.com/explore/{note_id}" if note_id else ''
                
                row = {
                    '笔记 ID': note_id,
                    '笔记链接': note_url,
                    '标题': note.get('title', note.get('display_title', '')),
                    '描述': note.get('desc', ''),
                    '作者': user.get('nickname', ''),
                    '作者 ID': user.get('userid', ''),
                    '点赞数': note.get('likes', 0),
                    '收藏数': note.get('collected_count', 0),
                    '评论数': note.get('comments_count', 0),
                    '分享数': note.get('share_count', 0),
                    '类型': note.get('type', ''),
                    '图片数量': len(images),
                    '封面图': cover_image_url,  # 实际图片链接
                    '发布时间': datetime.fromtimestamp(note.get('time', 0)).strftime('%Y-%m-%d %H:%M:%S') if note.get('time') else '',
                    'IP 属地': note.get('ip_location', '')
                }
                data_rows.append(row)
            
            # 创建 DataFrame
            df = pd.DataFrame(data_rows)
            
            # 导出到 Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='笔记详情')
                
                # 获取 workbook 和 worksheet
                workbook = writer.book
                worksheet = writer.sheets['笔记详情']
                
                # 设置列宽自适应和自动换行
                for col_idx, column in enumerate(df.columns, 1):
                    # 计算列的最大宽度
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    
                    # 获取列标题长度
                    header_length = len(str(column))
                    max_length = max(max_length, header_length)
                    
                    # 获取该列所有单元格的最大长度
                    for row_idx in range(2, len(df) + 2):
                        cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            cell_length = len(str(cell_value))
                            max_length = min(max(max_length, cell_length), 50)  # 最大宽度 50
                    
                    # 设置列宽
                    worksheet.column_dimensions[column_letter].width = max_length + 2
                    
                    # 设置自动换行（对文本较长的列）
                    for row_idx in range(1, len(df) + 2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        # 对标题、描述、封面图等列设置自动换行
                        if column in ['标题', '描述', '封面图', '作者']:
                            cell.alignment = Alignment(wrap_text=True)
            
            return True
        
        except Exception as e:
            print(f"❌ 导出失败：{e}")
            return False


def main():
    """主函数 - 示例用法"""
    api_token = input("请输入您的 API Token: ").strip()
    
    if not api_token:
        print("❌ API Token 不能为空！")
        return
    
    crawler = XHSNoteDetail(api_token)
    
    print("\n" + "=" * 80)
    print("小红书笔记详情采集系统")
    print("=" * 80)
    
    print("\n请选择采集模式：")
    print("1. 单个笔记详情")
    print("2. 批量笔记详情")
    
    mode = input("\n请选择 (1-2, 默认 1): ").strip()
    
    if mode == "2":
        # 批量采集
        note_ids_input = input("\n请输入笔记 ID 列表（用逗号分隔）: ").strip()
        note_ids = [nid.strip() for nid in note_ids_input.split(",") if nid.strip()]
        
        if not note_ids:
            print("❌ 至少需要一个笔记 ID！")
            return
        
        print(f"\n🔍 正在获取 {len(note_ids)} 条笔记详情...")
        print("请稍候...\n")
        
        result = crawler.get_notes_by_ids(note_ids, delay_seconds=1.0)
        
        if result.get("success"):
            notes = result.get("data", {}).get("notes", [])
            print(f"✅ 成功获取 {len(notes)}/{len(note_ids)} 条笔记\n")
            
            for i, note in enumerate(notes, 1):
                temp_result = {"success": True, "data": {"note": note}}
                print(crawler.format_note_detail(temp_result))
        else:
            print(f"❌ 获取失败：{result.get('message')}")
    
    else:
        # 单个笔记
        note_id = input("\n请输入笔记 ID: ").strip()
        if not note_id:
            print("❌ 笔记 ID 不能为空！")
            return
        
        print(f"\n🔍 正在获取笔记 {note_id} 的详情...\n")
        
        result = crawler.get_note_detail(note_id)
        
        print(crawler.format_note_detail(result))
        
        # 询问是否保存
        if result.get("success"):
            save = input("\n是否保存结果为 JSON 文件？(y/n): ").strip().lower()
            if save == 'y':
                filename = f"note_{note_id}_detail.json"
                if crawler.export_to_json(result, filename):
                    print(f"✅ 已保存到：{filename}")
                else:
                    print("❌ 保存失败")


if __name__ == "__main__":
    main()
