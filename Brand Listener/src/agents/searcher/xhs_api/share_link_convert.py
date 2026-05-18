import requests
import gzip
import json
import re
from typing import Dict, Any, List, Optional
from datetime import datetime
from pathlib import Path


class XHSShareLinkConverter:
    """小红书分享链接转化类"""
    
    def __init__(self, api_token: str):
        """
        初始化分享链接转化类
        
        Args:
            api_token: API 访问令牌
        """
        self.base_url = "https://proxy-api.ainm.store"
        self.api_token = api_token
        self.endpoint = "/p/xiaohongshu/share-url-transfer/v1"
    
    def convert_share_url(
        self,
        share_url: str
    ) -> Dict[str, Any]:
        """
        将分享短链接转换为标准笔记链接
        
        Args:
            share_url: 小红书分享链接（短链接或普通分享链接）
        
        Returns:
            包含转换结果的字典
        """
        # 先尝试直接从链接提取笔记 ID（适用于标准格式链接）
        note_id = self._extract_note_id_from_url(share_url)
        
        if note_id:
            # 已经是标准格式，直接返回
            return {
                "success": True,
                "data": {
                    "redirect_url": share_url,
                    "note_id": note_id
                },
                "message": "链接已是标准格式"
            }
        
        # 如果不是标准格式，调用 API 转换
        url = f"{self.base_url}{self.endpoint}"
        
        headers = {
            "Authorization": f"Bearer {self.api_token}"
        }
        
        params = {
            "shareUrl": share_url
        }
        
        try:
            response = requests.get(url, params=params, headers=headers, timeout=60)
            response.raise_for_status()
            
            # 解析响应（支持 gzip 压缩）
            result = self._parse_response(response)
            
            # 处理嵌套的 code 结构：外层 code=200，内层 code=0
            outer_code = result.get("code") if result else None
            inner_data = result.get("data", {}) if result else {}
            inner_code = inner_data.get("code") if isinstance(inner_data, dict) else None
            
            # 检查是否成功（外层 200 且内层 0）
            if outer_code == 200 and inner_code == 0:
                redirect_url = inner_data.get("data", {}).get("redirect_url", "")
                
                # 从重定向链接提取笔记 ID
                note_id = self._extract_note_id_from_url(redirect_url)
                
                return {
                    "success": True,
                    "data": {
                        "redirect_url": redirect_url,
                        "note_id": note_id
                    },
                    "message": "转换成功"
                }
            else:
                # 获取错误信息（优先使用内层 message）
                error_message = None
                if isinstance(inner_data, dict):
                    error_message = inner_data.get("message")
                if not error_message:
                    error_message = self._get_error_message(outer_code) if outer_code else "响应解析失败"
                
                return {
                    "success": False,
                    "error_code": inner_code if inner_code else outer_code if outer_code else "UNKNOWN",
                    "message": error_message
                }
        
        except requests.exceptions.Timeout:
            return {
                "success": False,
                "error_code": "TIMEOUT",
                "message": "请求超时，建议重试"
            }
        except requests.exceptions.RequestException as e:
            return {
                "success": False,
                "error_code": "NETWORK_ERROR",
                "message": f"网络错误：{str(e)}"
            }
        except json.JSONDecodeError:
            return {
                "success": False,
                "error_code": "PARSE_ERROR",
                "message": "响应解析失败"
            }
    
    def _extract_note_id_from_url(self, url: str) -> str:
        """从标准笔记链接提取笔记 ID"""
        # 匹配 /discovery/item/{note_id} 或 /explore/{note_id}
        patterns = [
            r'/discovery/item/([a-f0-9]+)',
            r'/explore/([a-f0-9]+)',
            r'item/([a-f0-9]+)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        
        return ""
    
    def _get_error_message(self, code: int) -> str:
        """获取错误码对应的中文说明"""
        error_messages = {
            100: "Token 无效或已失效",
            301: "采集失败，请重试",
            302: "超出速率限制",
            303: "超出每日配额",
            400: "参数错误",
            500: "内部服务器错误",
            600: "权限不足",
            601: "余额不足"
        }
        return error_messages.get(code, f"未知错误 (code: {code})")
    
    @staticmethod
    def _parse_response(response) -> Dict[str, Any]:
        """解析响应，支持 gzip 压缩"""
        raw = response.content
        
        try:
            # 尝试 gzip 解压
            decompressed = gzip.decompress(raw)
            return json.loads(decompressed.decode('utf-8'))
        except:
            # 尝试直接解码
            try:
                return json.loads(raw.decode('utf-8'))
            except:
                return None
    
    def format_result(self, result: Dict[str, Any]) -> str:
        """格式化输出转换结果"""
        if not result.get("success"):
            return f"❌ 转换失败：{result.get('message')}"
        
        data = result.get("data", {})
        redirect_url = data.get("redirect_url", "")
        note_id = data.get("note_id", "")
        
        output = []
        output.append("=" * 80)
        output.append("✅ 分享链接转换成功")
        output.append("=" * 80)
        output.append(f"\n🔗 原始链接：{redirect_url[:100]}...")
        output.append(f"🆔 笔记 ID: {note_id}")
        output.append(f"\n💡 提示：可以使用该笔记 ID 调用笔记详情或评论 API")
        output.append("=" * 80)
        
        return "\n".join(output)
    
    def convert_batch(
        self,
        share_urls: List[str],
        delay_seconds: float = 0.5
    ) -> Dict[str, Any]:
        """
        批量转换分享链接
        
        Args:
            share_urls: 分享链接列表
            delay_seconds: 每个链接之间的延迟时间（秒）
        
        Returns:
            包含所有转换结果的字典
        """
        import time
        
        results = []
        success_count = 0
        failed_count = 0
        
        for i, url in enumerate(share_urls, 1):
            print(f"\n处理第 {i}/{len(share_urls)} 个链接...")
            
            result = self.convert_share_url(url)
            
            if result.get("success"):
                success_count += 1
                print(f"✅ 转换成功：笔记 ID = {result.get('data', {}).get('note_id', '')}")
            else:
                failed_count += 1
                print(f"❌ 转换失败：{result.get('message')}")
            
            results.append({
                "original_url": url,
                "result": result
            })
            
            # 延迟
            if i < len(share_urls):
                time.sleep(delay_seconds)
        
        return {
            "success": True,
            "data": {
                "results": results,
                "total": len(share_urls),
                "success_count": success_count,
                "failed_count": failed_count
            },
            "message": f"批量转换完成：成功 {success_count} 个，失败 {failed_count} 个"
        }
    
    def export_to_excel(
        self,
        batch_result: Dict[str, Any],
        filename: str
    ) -> bool:
        """
        导出批量转换结果到 Excel 文件
        
        Args:
            batch_result: 批量转换结果
            filename: 输出文件名
        
        Returns:
            是否导出成功
        """
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment, PatternFill, Font
            
            if not batch_result.get("success"):
                return False
            
            data = batch_result.get("data", {})
            results = data.get("results", [])
            
            if not results:
                return False
            
            # 提取字段
            data_rows = []
            for item in results:
                original_url = item.get("original_url", "")
                result = item.get("result", {})
                
                if result.get("success"):
                    result_data = result.get("data", {})
                    row = {
                        '序号': len(data_rows) + 1,
                        '状态': '成功',
                        '原始链接': original_url,
                        '转换后链接': result_data.get("redirect_url", ""),
                        '笔记 ID': result_data.get("note_id", ""),
                        '消息': result.get("message", "")
                    }
                else:
                    row = {
                        '序号': len(data_rows) + 1,
                        '状态': '失败',
                        '原始链接': original_url,
                        '转换后链接': '',
                        '笔记 ID': '',
                        '消息': result.get("message", "")
                    }
                
                data_rows.append(row)
            
            # 创建 DataFrame
            df = pd.DataFrame(data_rows)
            
            # 导出到 Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='链接转换结果')
                
                # 获取 workbook 和 worksheet
                workbook = writer.book
                worksheet = writer.sheets['链接转换结果']
                
                # 设置列宽自适应和自动换行
                for col_idx, column in enumerate(df.columns, 1):
                    # 计算列的最大宽度
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    
                    # 获取列标题长度
                    header_length = len(str(column))
                    max_length = max(max_length, header_length)
                    
                    # 获取该列所有单元格的最大长度
                    for row_idx in range(2, len(df) + 2):
                        cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            cell_length = len(str(cell_value))
                            max_length = min(max(max_length, cell_length), 80)  # 最大宽度 80
                    
                    # 设置列宽
                    worksheet.column_dimensions[column_letter].width = max_length + 2
                    
                    # 设置自动换行（对链接列）
                    if column in ['原始链接', '转换后链接']:
                        for row_idx in range(1, len(df) + 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = Alignment(wrap_text=True)
                
                # 设置状态列的背景色
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=2)  # 状态在第 2 列
                    if cell.value == '成功':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True)
                    elif cell.value == '失败':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(bold=True)
            
            print(f"✅ 已导出 Excel: {filename}")
            print(f"   - 共 {len(data_rows)} 条记录（成功：{data['success_count']} 条，失败：{data['failed_count']} 条）")
            print(f"   - 状态标记：成功=绿色背景，失败=红色背景")
            return True
        
        except Exception as e:
            print(f"❌ 导出失败：{e}")
            return False


def main():
    """主函数 - 示例用法"""
    api_token = input("请输入您的 API Token: ").strip()
    
    if not api_token:
        print("❌ API Token 不能为空！")
        return
    
    converter = XHSShareLinkConverter(api_token)
    
    print("\n" + "=" * 80)
    print("小红书分享链接转化系统")
    print("=" * 80)
    
    share_url = input("\n请输入小红书分享链接：").strip()
    if not share_url:
        print("❌ 分享链接不能为空！")
        return
    
    print(f"\n🔍 正在转换链接...\n")
    
    result = converter.convert_share_url(share_url)
    
    print(converter.format_result(result))


if __name__ == "__main__":
    main()
