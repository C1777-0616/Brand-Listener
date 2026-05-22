import requests
import gzip
import json
from typing import Optional, Dict, Any, List
from datetime import datetime


class XHSNoteComments:
    """小红书笔记评论采集类"""
    
    def __init__(self, api_token: str):
        """
        初始化评论采集类
        
        Args:
            api_token: API 访问令牌
        """
        self.base_url = "https://proxy-api.ainm.store"
        self.api_token = api_token
        self.endpoint = "/p/xiaohongshu/get-note-comment/v2"
        # 子评论接口
        self.sub_comment_endpoint = "/p/xiaohongshu/get-note-sub-comment/v2"
    
    @staticmethod
    def _extract_cursor(cursor_value: Optional[str]) -> Optional[str]:
        """
        提取 cursor 的实际值
        
        API 返回的 cursor 可能是嵌套 JSON 格式：
        {"cursor":"actual_cursor_value","index":2,"pageArea":"UNFOLDED"}
        
        需要提取内层的 actual_cursor_value 作为下一页的 lastCursor
        
        Args:
            cursor_value: cursor 字符串
            
        Returns:
            提取后的 cursor 值
        """
        if not cursor_value:
            return None
        
        try:
            # 尝试解析为 JSON
            cursor_json = json.loads(cursor_value) if isinstance(cursor_value, str) else cursor_value
            # 如果解析成功且有 cursor 字段，返回内层 cursor
            if isinstance(cursor_json, dict) and "cursor" in cursor_json:
                return cursor_json.get("cursor")
        except (json.JSONDecodeError, TypeError):
            # 解析失败，说明已经是纯 cursor 字符串，直接返回
            pass
        
        # 返回原始值
        return cursor_value
    
    def get_comments(
        self,
        note_id: str,
        last_cursor: Optional[str] = None,
        sort: str = "latest",
        max_retries: int = 3
    ) -> Dict[str, Any]:
        """
        获取笔记评论列表
        
        Args:
            note_id: 笔记 ID（必填）
            last_cursor: 分页游标（可选）
            sort: 排序方式，默认 latest
                - normal: 默认排序
                - latest: 最新优先
            max_retries: 最大重试次数（针对 301 错误）
        
        Returns:
            包含评论列表的字典
        """
        url = f"{self.base_url}{self.endpoint}"
        
        headers = {
            "Authorization": f"Bearer {self.api_token}"
        }
        
        params = {
            "noteId": note_id,
            "sort": sort
        }
        
        if last_cursor:
            params["lastCursor"] = last_cursor
        
        # 重试逻辑
        import time
        
        for attempt in range(1, max_retries + 1):
            try:
                response = requests.get(url, params=params, headers=headers, timeout=60)
                response.raise_for_status()
                
                # 解析响应（支持 gzip 压缩）
                result = self._parse_response(response)
                
                if not result:
                    if attempt < max_retries:
                        print(f"   响应解析失败，{max_retries - attempt}秒后重试...")
                        time.sleep(max_retries - attempt)
                        continue
                    return {
                        "success": False,
                        "error_code": "PARSE_ERROR",
                        "message": "响应解析失败"
                    }
                
                # 修复：API 返回 code: 200 表示成功，不是 0
                if result and result.get("code") in [0, 200]:
                    # 处理嵌套的 data 结构
                    outer_data = result.get("data", {})
                    
                    # 如果 outer_data 是 dict 且包含 code 和 data 字段，说明是嵌套结构
                    if isinstance(outer_data, dict) and "code" in outer_data and "data" in outer_data:
                        # 使用内层的 data
                        inner_data = outer_data.get("data", {})
                    else:
                        # 已经是 flat 结构，直接使用
                        inner_data = outer_data
                    
                    return {
                        "success": True,
                        "data": inner_data,
                        "message": "获取成功"
                    }
                elif result.get("code") == 301:
                    if attempt < max_retries:
                        wait_time = attempt * 5
                        print(f"   301 错误，等待 {wait_time}秒后重试 ({attempt}/{max_retries})...")
                        time.sleep(wait_time)
                        continue
                    else:
                        return {
                            "success": False,
                            "error_code": 301,
                            "message": "采集失败，请重试"
                        }
                else:
                    return {
                        "success": False,
                        "error_code": result.get("code"),
                        "message": self._get_error_message(result.get("code"))
                    }
            
            except requests.exceptions.Timeout:
                if attempt < max_retries:
                    print(f"   请求超时，{max_retries - attempt}秒后重试...")
                    time.sleep(max_retries - attempt)
                    continue
                return {
                    "success": False,
                    "error_code": "TIMEOUT",
                    "message": "请求超时，建议重试"
                }
            except requests.exceptions.RequestException as e:
                if attempt < max_retries:
                    print(f"   网络错误，{max_retries - attempt}秒后重试...")
                    time.sleep(max_retries - attempt)
                    continue
                return {
                    "success": False,
                    "error_code": "NETWORK_ERROR",
                    "message": f"网络错误：{str(e)}"
                }
        
        # 不应该到这里
        return {
            "success": False,
            "error_code": "UNKNOWN",
            "message": "未知错误"
        }
    
    def get_all_comments(
        self,
        note_id: str,
        max_pages: Optional[int] = None,
        delay_seconds: float = 1.0,
        sort: str = "latest",
        fetch_sub_comments: bool = False,
        sub_comment_delay: float = 0.5
    ) -> Dict[str, Any]:
        """
        获取笔记的所有评论（自动分页）
        
        Args:
            note_id: 笔记 ID
            max_pages: 最大页数（可选，不限制则为 None）
            delay_seconds: 每页之间的延迟时间（秒），默认 1 秒
            sort: 排序方式
            fetch_sub_comments: 是否采集子评论（回复），默认 False
            sub_comment_delay: 每条子评论采集的延迟时间（秒），默认 0.5 秒
        
        Returns:
            包含所有评论的字典
        """
        import time
        
        all_comments = []
        last_cursor = None
        page = 1
        has_more = True
        
        while has_more:
            if max_pages and page > max_pages:
                break
            
            result = self.get_comments(note_id, last_cursor, sort)
            
            if not result.get("success"):
                # 第 1 页失败才报错；后续页失败返回已有结果
                if page == 1:
                    return {
                        "success": False,
                        "error_code": result.get("error_code"),
                        "message": f"第 {page} 页获取失败：{result.get('message')}",
                        "data": {"comments": all_comments, "total_pages": page - 1}
                    }
                else:
                    print(f"[WARN] 第 {page} 页获取失败，返回已有 {len(all_comments)} 条评论")
                    break
            
            data = result.get("data", {})
            comments = data.get("comments", [])
            
            # 如果需要采集子评论
            if fetch_sub_comments:
                for comment in comments:
                    self._fetch_sub_comments(comment, sub_comment_delay)
            
            all_comments.extend(comments)
            
            print(f"[OK] 已获取第 {page} 页，共 {len(comments)} 条评论")
            
            has_more = data.get("has_more", False)
            
            if has_more and comments:
                # 获取 cursor 并提取实际值
                raw_cursor = data.get("cursor")
                last_cursor = self._extract_cursor(raw_cursor)
                
                if not last_cursor:
                    break
                
                page += 1
                
                # 延迟以避免触发限流
                if has_more and (not max_pages or page <= max_pages):
                    time.sleep(delay_seconds)
            else:
                break
        
        return {
            "success": True,
            "data": {
                "comments": all_comments,
                "total_pages": page,
                "total_count": len(all_comments),
                "comment_count": data.get("comment_count", 0)
            },
            "message": f"获取成功，共 {len(all_comments)} 条评论"
        }
    
    def _fetch_sub_comments(
        self,
        parent_comment: Dict[str, Any],
        delay_seconds: float = 0.5
    ) -> Dict[str, Any]:
        """
        获取单条评论的所有子评论（回复）
        使用专门的子评论接口：/api/xiaohongshu/get-note-sub-comment/v2
        
        核心策略：
        1. 主评论接口返回的 sub_comments 只包含第 1 条子评论
        2. 使用子评论接口 + lastCursor 分页获取所有子评论
        3. 合并时去重
        
        Args:
            parent_comment: 父评论
            delay_seconds: 请求延迟时间
        
        Returns:
            包含子评论的评论对象
        """
        import time
        
        sub_comment_count = parent_comment.get("sub_comment_count", 0)
        
        # 如果没有子评论，直接返回
        if sub_comment_count == 0:
            return parent_comment
        
        # 从主评论响应中获取已有的子评论（第 1 条）
        existing_sub_comments = parent_comment.get("sub_comments", [])
        existing_ids = {c.get("id") for c in existing_sub_comments}
        
        # 如果已经有所有子评论，直接返回
        if len(existing_sub_comments) >= sub_comment_count:
            parent_comment["sub_comments"] = existing_sub_comments
            return parent_comment
        
        # 使用子评论接口获取所有子评论（包括第 1 条）
        all_sub_comments = []
        last_cursor = None  # 第一次请求不需要 cursor
        
        while True:
            # 使用专门的子评论接口
            url = f"{self.base_url}{self.sub_comment_endpoint}"
            
            headers = {
                "Authorization": f"Bearer {self.api_token}"
            }
            
            params = {
                "noteId": parent_comment.get("note_id"),
                "commentId": parent_comment.get("id"),
            }
            
            # 添加分页 cursor（第一次为 None）
            if last_cursor:
                params["lastCursor"] = last_cursor
            
            try:
                response = requests.get(url, params=params, headers=headers, timeout=60)
                response.raise_for_status()
                
                # 解析响应（支持 gzip 压缩）
                result = self._parse_response(response)
                
                # 处理嵌套的 code 结构：外层 200，内层 0
                outer_code = result.get("code") if result else None
                inner_data = result.get("data", {}) if result else {}
                inner_code = inner_data.get("code") if isinstance(inner_data, dict) else None
                
                if not result or (outer_code != 200 and inner_code != 0):
                    print(f"[WARN] 获取子评论失败：{result.get('message') if result else '响应解析失败'}")
                    break
                
                # 处理嵌套的 data 结构
                if isinstance(inner_data, dict) and "data" in inner_data:
                    data = inner_data.get("data", {})
                else:
                    data = inner_data
                new_sub_comments = data.get("comments", [])
                
                if not new_sub_comments:
                    break
                
                # 去重：只添加之前没有的子评论
                for sub in new_sub_comments:
                    if sub.get("id") not in existing_ids:
                        all_sub_comments.append(sub)
                        existing_ids.add(sub.get("id"))
                
                # 获取下一页 cursor
                raw_cursor = data.get("cursor")
                last_cursor = self._extract_cursor(raw_cursor)
                
                # 没有更多页面，退出循环
                if not last_cursor or not data.get("has_more", False):
                    break
                
                # 延迟
                time.sleep(delay_seconds)
                
            except requests.exceptions.Timeout:
                print("[WARN] 获取子评论超时")
                break
            except Exception as e:
                print(f"[WARN] 获取子评论错误：{e}")
                break
        
        # 合并子评论列表：已有的 + 新获取的
        all_sub_comments = existing_sub_comments + all_sub_comments
        parent_comment["sub_comments"] = all_sub_comments
        return parent_comment
    
    def _get_error_message(self, code: int) -> str:
        """获取错误码对应的中文说明"""
        error_messages = {
            100: "Token 无效或已失效",
            301: "采集失败，请重试",
            302: "超出速率限制",
            303: "超出每日配额",
            400: "参数错误",
            500: "内部服务器错误",
            600: "权限不足",
            601: "余额不足",
            6002: "评论接口异常，请稍后重试"
        }
        return error_messages.get(code, f"未知错误 (code: {code})")
    
    @staticmethod
    def _parse_response(response) -> Optional[Dict[str, Any]]:
        """解析响应，支持 gzip 压缩"""
        raw = response.content
        
        try:
            # 尝试 gzip 解压
            decompressed = gzip.decompress(raw)
            return json.loads(decompressed.decode('utf-8'))
        except:
            # 尝试直接解码
            try:
                return json.loads(raw.decode('utf-8'))
            except:
                return None
    
    def format_comments(self, comments_result: Dict[str, Any]) -> str:
        """格式化输出评论列表"""
        if not comments_result.get("success"):
            return f"❌ 获取失败：{comments_result.get('message')}"
        
        data = comments_result.get("data", {})
        comments = data.get("comments", [])
        
        if not comments:
            return "ℹ️ 该笔记暂无评论"
        
        output = []
        total = data.get("total_count", len(comments))
        output.append(f"✅ 获取成功，共 {total} 条评论\n")
        output.append("=" * 80)
        
        for idx, comment in enumerate(comments, 1):
            user = comment.get("user", {})
            
            output.append(f"\n【{idx}】{comment.get('id', 'N/A')}")
            output.append(f"内容：{comment.get('content', 'N/A')[:100]}")
            output.append(f"用户：{user.get('nickname', 'N/A')} (ID: {user.get('userid', 'N/A')})")
            
            if user.get("red_id"):
                output.append(f"小红书号：{user.get('red_id')}")
            
            output.append(f"点赞：{comment.get('like_count', 0)}")
            
            # IP 位置
            if comment.get("ip_location"):
                output.append(f"IP 位置：{comment.get('ip_location')}")
            
            # 子评论数
            sub_count = comment.get("sub_comment_count", 0)
            if sub_count > 0:
                output.append(f"子评论：{sub_count} 条")
            
            # 发布时间
            time_val = comment.get("time")
            if time_val:
                dt = datetime.fromtimestamp(time_val)
                output.append(f"发布时间：{dt.strftime('%Y-%m-%d %H:%M:%S')}")
            
            # 标签
            show_tags = comment.get("show_tags_v2", [])
            if show_tags:
                tags = [tag.get("text", "") for tag in show_tags if tag.get("text")]
                if tags:
                    output.append(f"标签：{', '.join(tags)}")
            
            output.append("-" * 80)
        
        return "\n".join(output)
    
    def export_to_json(self, comments_result: Dict[str, Any], filename: str) -> bool:
        """
        导出评论数据到 JSON 文件
        
        Args:
            comments_result: 评论数据结果
            filename: 输出文件名
        
        Returns:
            是否导出成功
        """
        if not comments_result.get("success"):
            return False
        
        try:
            data = comments_result.get("data", {})
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return True
        except Exception:
            return False
    
    def export_to_excel(
        self,
        comments_result: Dict[str, Any],
        filename: str,
        include_sub_comments: bool = True
    ) -> bool:
        """
        导出评论数据到 Excel 文件（主评论和子评论汇总到一张表）
        
        Args:
            comments_result: 评论数据结果
            filename: 输出文件名
            include_sub_comments: 是否包含子评论，默认 True
        
        Returns:
            是否导出成功
        """
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment, PatternFill, Font
            
            if not comments_result.get("success"):
                return False
            
            data = comments_result.get("data", {})
            comments = data.get("comments", [])
            
            if not comments:
                return False
            
            # 提取字段（带全局去重）
            data_rows = []
            row_num = 1
            seen_comment_ids = set()  # 全局去重
            
            for comment in comments:
                # 主评论去重
                comment_id = comment.get('id', '')
                if comment_id in seen_comment_ids:
                    continue
                seen_comment_ids.add(comment_id)
                
                user = comment.get("user", {})
                show_tags = comment.get("show_tags_v2", [])
                
                # 提取标签文本
                tag_texts = []
                if show_tags:
                    tag_texts = [tag.get("text", "") for tag in show_tags if tag.get("text")]
                
                # 主评论
                row = {
                    '序号': row_num,
                    '评论类型': '主评论',
                    '评论 ID': comment_id,
                    '内容': comment.get('content', ''),
                    '用户昵称': user.get('nickname', ''),
                    '用户 ID': user.get('userid', ''),
                    '小红书号': user.get('red_id', ''),
                    '点赞数': comment.get('like_count', 0),
                    'IP 属地': comment.get('ip_location', ''),
                    '标签': ', '.join(tag_texts),
                    '发布时间': datetime.fromtimestamp(comment.get('time', 0)).strftime('%Y-%m-%d %H:%M:%S') if comment.get('time') else '',
                    '回复对象': '',
                    '父评论 ID': ''
                }
                data_rows.append(row)
                row_num += 1
                
                # 如果需要包含子评论
                if include_sub_comments:
                    sub_comments = comment.get("sub_comments", [])
                    for sub_comment in sub_comments:
                        # 子评论去重
                        sub_id = sub_comment.get('id', '')
                        if sub_id in seen_comment_ids:
                            continue
                        seen_comment_ids.add(sub_id)
                        sub_user = sub_comment.get("user", {})
                        sub_tags = sub_comment.get("show_tags_v2", [])
                        
                        # 提取子评论标签
                        sub_tag_texts = []
                        if sub_tags:
                            sub_tag_texts = [tag.get("text", "") for tag in sub_tags if tag.get("text")]
                        
                        # 检查是否有回复目标
                        target_comment = sub_comment.get("target_comment", {})
                        target_user = target_comment.get("user", {})
                        reply_to = target_user.get("nickname", "")
                        
                        sub_row = {
                            '序号': row_num,
                            '评论类型': '子评论',
                            '评论 ID': sub_comment.get('id', ''),
                            '内容': sub_comment.get('content', ''),
                            '用户昵称': sub_user.get('nickname', ''),
                            '用户 ID': sub_user.get('userid', ''),
                            '小红书号': sub_user.get('red_id', ''),
                            '点赞数': sub_comment.get('like_count', 0),
                            'IP 属地': sub_comment.get('ip_location', ''),
                            '标签': ', '.join(sub_tag_texts),
                            '发布时间': datetime.fromtimestamp(sub_comment.get('time', 0)).strftime('%Y-%m-%d %H:%M:%S') if sub_comment.get('time') else '',
                            '回复对象': reply_to,
                            '父评论 ID': comment.get('id', '')
                        }
                        data_rows.append(sub_row)
                        row_num += 1
            
            # 创建 DataFrame
            df = pd.DataFrame(data_rows)
            
            # 导出到 Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='评论数据')
                
                # 获取 workbook 和 worksheet
                workbook = writer.book
                worksheet = writer.sheets['评论数据']
                
                # 设置列宽自适应和自动换行
                for col_idx, column in enumerate(df.columns, 1):
                    # 计算列的最大宽度
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    
                    # 获取列标题长度
                    header_length = len(str(column))
                    max_length = max(max_length, header_length)
                    
                    # 获取该列所有单元格的最大长度
                    for row_idx in range(2, len(df) + 2):
                        cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            cell_length = len(str(cell_value))
                            max_length = min(max(max_length, cell_length), 80)  # 最大宽度 80
                    
                    # 设置列宽
                    worksheet.column_dimensions[column_letter].width = max_length + 2
                    
                    # 设置自动换行（对文本较长的列）
                    for row_idx in range(1, len(df) + 2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        # 对内容、用户昵称等列设置自动换行
                        if column in ['内容', '用户昵称', '标签', '回复对象']:
                            cell.alignment = Alignment(wrap_text=True)
                
                # 设置评论类型列的背景色
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=2)  # 评论类型在第 2 列
                    if cell.value == '主评论':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True)
                    elif cell.value == '子评论':
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            print(f"[OK] 已导出 Excel: {filename}")
            print(f"   - 共 {len(data_rows)} 条评论（主评论：{len(comments)} 条，子评论：{len(data_rows) - len(comments)} 条）")
            print(f"   - 评论类型已标记（主评论：绿色背景，子评论：黄色背景）")
            return True
        
        except Exception as e:
            print(f"[FAIL] 导出失败：{e}")
            return False
    
    def export_to_csv(self, comments_result: Dict[str, Any], filename: str) -> bool:
        """
        导出评论数据到 CSV 文件
        
        Args:
            comments_result: 评论数据结果
            filename: 输出文件名
        
        Returns:
            是否导出成功
        """
        if not comments_result.get("success"):
            return False
        
        try:
            import csv
            
            data = comments_result.get("data", {})
            comments = data.get("comments", [])
            
            if not comments:
                return False
            
            # CSV 字段
            fieldnames = [
                'comment_id', 'content', 'like_count', 'sub_comment_count',
                'create_time', 'ip_location', 'user_id', 'nickname', 'red_id'
            ]
            
            with open(filename, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                
                for comment in comments:
                    user = comment.get("user", {})
                    
                    row = {
                        'comment_id': comment.get('id', ''),
                        'content': comment.get('content', ''),
                        'like_count': comment.get('like_count', 0),
                        'sub_comment_count': comment.get('sub_comment_count', 0),
                        'create_time': comment.get('time', ''),
                        'ip_location': comment.get('ip_location', ''),
                        'user_id': user.get('userid', ''),
                        'nickname': user.get('nickname', ''),
                        'red_id': user.get('red_id', '')
                    }
                    writer.writerow(row)
            
            return True
        except Exception:
            return False


def main():
    """主函数 - 示例用法"""
    api_token = input("请输入您的 API Token: ").strip()
    
    if not api_token:
        print("❌ API Token 不能为空！")
        return
    
    crawler = XHSNoteComments(api_token)
    
    print("\n" + "=" * 80)
    print("小红书笔记评论采集系统")
    print("=" * 80)
    
    note_id = input("\n请输入笔记 ID: ").strip()
    if not note_id:
        print("❌ 笔记 ID 不能为空！")
        return
    
    print("\n选择采集模式：")
    print("1. 单页采集（仅第一页）")
    print("2. 全量采集（所有评论，自动分页）")
    
    mode = input("\n请选择 (1-2, 默认 1): ").strip()
    
    print("\n选择排序方式：")
    print("1. 最新优先 (latest)")
    print("2. 默认排序 (normal)")
    
    sort_choice = input("\n请选择 (1-2, 默认 1): ").strip()
    sort = "latest" if sort_choice != "2" else "normal"
    
    if mode == "2":
        max_pages_input = input("请输入最大页数（留空表示不限制）: ").strip()
        max_pages = int(max_pages_input) if max_pages_input else None
        
        delay_input = input("请输入每页延迟时间（秒，默认 1.0）: ").strip()
        delay = float(delay_input) if delay_input else 1.0
        
        print(f"\n🔍 正在采集笔记 {note_id} 的所有评论...")
        print("请稍候，这可能需要一些时间...\n")
        
        result = crawler.get_all_comments(note_id, max_pages, delay, sort)
    else:
        print(f"\n🔍 正在采集笔记 {note_id} 的评论（第一页）...\n")
        result = crawler.get_comments(note_id, sort=sort)
    
    print(crawler.format_comments(result))
    
    # 询问是否导出
    if result.get("success"):
        export = input("\n是否导出数据？(y/n): ").strip().lower()
        if export == 'y':
            format_type = input("选择导出格式 (json/csv): ").strip().lower()
            if format_type == 'json':
                filename = f"note_{note_id}_comments.json"
                if crawler.export_to_json(result, filename):
                    print(f"✅ 已导出到：{filename}")
                else:
                    print("❌ 导出失败")
            elif format_type == 'csv':
                filename = f"note_{note_id}_comments.csv"
                if crawler.export_to_csv(result, filename):
                    print(f"✅ 已导出到：{filename}")
                else:
                    print("❌ 导出失败")
            else:
                print("❌ 不支持的格式")


if __name__ == "__main__":
    main()
