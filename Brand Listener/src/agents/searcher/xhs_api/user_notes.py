import requests
import gzip
import json
from typing import Optional, Dict, Any, List
from datetime import datetime


class XHSUserNotes:
    """小红书用户笔记采集类"""
    
    def __init__(self, api_token: str):
        """
        初始化用户笔记采集类
        
        Args:
            api_token: API 访问令牌
        """
        self.base_url = "https://proxy-api.ainm.store"
        self.api_token = api_token
        self.endpoint = "/p/xiaohongshu/get-user-note-list/v4"
    
    def get_user_notes(
        self,
        user_id: str,
        last_cursor: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        获取用户发布的笔记列表
        
        Args:
            user_id: 小红书用户的唯一标识符（必填）
            last_cursor: 上一页的分页游标（可选，用于分页）
        
        Returns:
            包含用户笔记列表的字典
        """
        url = f"{self.base_url}{self.endpoint}"
        
        headers = {
            "Authorization": f"Bearer {self.api_token}"
        }
        
        params = {
            "userId": user_id
        }
        
        if last_cursor:
            params["lastCursor"] = last_cursor
        
        try:
            response = requests.get(url, params=params, headers=headers, timeout=60)
            response.raise_for_status()
            
            # 解析响应（支持 gzip 压缩）
            result = self._parse_response(response)
            
            # 修复：API 返回 code: 200 表示成功，不是 0
            if result and result.get("code") in [0, 200]:
                # 修复：处理嵌套的 data 字段
                outer_data = result.get("data", {})
                # 如果有嵌套的 data 字段，使用内层 data
                if isinstance(outer_data, dict) and "data" in outer_data:
                    inner_data = outer_data.get("data", {})
                else:
                    inner_data = outer_data
                
                return {
                    "success": True,
                    "data": inner_data,
                    "message": "获取成功"
                }
            elif result:
                return {
                    "success": False,
                    "error_code": result.get("code"),
                    "message": self._get_error_message(result.get("code"))
                }
            else:
                return {
                    "success": False,
                    "error_code": "PARSE_ERROR",
                    "message": "响应解析失败"
                }
        
        except requests.exceptions.Timeout:
            return {
                "success": False,
                "error_code": "TIMEOUT",
                "message": "请求超时，建议重试"
            }
        except requests.exceptions.RequestException as e:
            return {
                "success": False,
                "error_code": "NETWORK_ERROR",
                "message": f"网络错误：{str(e)}"
            }
    
    def _parse_response(self, response) -> Optional[Dict[str, Any]]:
        """解析响应，支持 gzip 压缩"""
        raw = response.content
        
        try:
            # 尝试 gzip 解压
            decompressed = gzip.decompress(raw)
            return json.loads(decompressed.decode('utf-8'))
        except:
            # 尝试直接解码
            try:
                return json.loads(raw.decode('utf-8'))
            except:
                return None
    
    @staticmethod
    def _extract_cursor(cursor_value: Optional[str]) -> Optional[str]:
        """
        提取 cursor 的实际值
        
        API 返回的 cursor 可能是嵌套 JSON 格式：
        {"cursor":"actual_cursor_value","index":2,"pageArea":"UNFOLDED"}
        
        需要提取内层的 actual_cursor_value 作为下一页的 lastCursor
        
        Args:
            cursor_value: cursor 字符串
        
        Returns:
            提取后的 cursor 值
        """
        if not cursor_value:
            return None
        
        try:
            # 尝试解析为 JSON
            cursor_json = json.loads(cursor_value) if isinstance(cursor_value, str) else cursor_value
            
            # 如果解析成功且有 cursor 字段，返回内层 cursor
            if isinstance(cursor_json, dict) and "cursor" in cursor_json:
                return cursor_json.get("cursor")
            
            # 解析失败，说明已经是纯 cursor 字符串，直接返回
            return cursor_value
        except:
            # 解析失败，直接返回原值
            return cursor_value
    
    def get_all_user_notes(
        self,
        user_id: str,
        max_pages: Optional[int] = None,
        delay_seconds: float = 1.0
    ) -> Dict[str, Any]:
        """
        获取用户的所有笔记（自动分页）
        
        Args:
            user_id: 小红书用户的唯一标识符
            max_pages: 最大页数（可选，不限制则为 None）
            delay_seconds: 每页之间的延迟时间（秒），默认 1 秒
        
        Returns:
            包含所有笔记的字典
        """
        import time
        
        all_notes = []
        last_cursor = None
        page = 1
        has_more = True
        
        while has_more:
            if max_pages and page > max_pages:
                break
            
            result = self.get_user_notes(user_id, last_cursor)
            
            if not result.get("success"):
                return {
                    "success": False,
                    "error_code": result.get("error_code"),
                    "message": f"第 {page} 页获取失败：{result.get('message')}",
                    "data": {"notes": all_notes, "total_pages": page - 1}
                }
            
            data = result.get("data", {})
            notes = data.get("notes", [])
            all_notes.extend(notes)
            
            print(f"✅ 已获取第 {page} 页，共 {len(notes)} 条笔记")
            
            has_more = data.get("has_more", False)
            
            if has_more and notes:
                # 获取最后一条笔记的 cursor 并提取实际值
                raw_cursor = notes[-1].get("cursor")
                last_cursor = self._extract_cursor(raw_cursor)
                
                if not last_cursor:
                    print(f"   ⚠️  未获取到分页游标，停止采集")
                    break
                
                page += 1
                
                # 延迟以避免触发限流
                if has_more and (not max_pages or page <= max_pages):
                    time.sleep(delay_seconds)
            else:
                break
        
        return {
            "success": True,
            "data": {
                "notes": all_notes,
                "total_pages": page,
                "total_count": len(all_notes)
            },
            "message": f"获取成功，共 {len(all_notes)} 条笔记"
        }
    
    def _get_error_message(self, code: int) -> str:
        """获取错误码对应的中文说明"""
        error_messages = {
            100: "Token 无效或已失效",
            301: "采集失败，请重试",
            302: "超出速率限制",
            303: "超出每日配额",
            400: "参数错误",
            500: "内部服务器错误",
            600: "权限不足",
            601: "余额不足"
        }
        return error_messages.get(code, f"未知错误 (code: {code})")
    
    def format_notes(self, notes_result: Dict[str, Any]) -> str:
        """格式化输出用户笔记列表"""
        if not notes_result.get("success"):
            return f"❌ 获取失败：{notes_result.get('message')}"
        
        data = notes_result.get("data", {})
        notes = data.get("notes", [])
        
        if not notes:
            return "ℹ️ 该用户暂无笔记"
        
        output = []
        total = data.get("total_count", len(notes))
        output.append(f"✅ 获取成功，共 {total} 条笔记\n")
        output.append("=" * 80)
        
        for idx, note in enumerate(notes, 1):
            user = note.get("user", {})
            
            output.append(f"\n【{idx}】{note.get('id', 'N/A')}")
            output.append(f"标题：{note.get('title', note.get('display_title', 'N/A'))[:100]}")
            output.append(f"描述：{note.get('desc', 'N/A')[:100]}...")
            output.append(f"作者：{user.get('nickname', 'N/A')} (ID: {user.get('userid', 'N/A')})")
            output.append(f"点赞：{note.get('likes', 0)} | 收藏：{note.get('collected_count', 0)} | 评论：{note.get('comments_count', 0)}")
            
            if note.get("images_list"):
                cover = note["images_list"][0].get("url", "")
                output.append(f"封面：{cover[:80]}...")
            
            output.append(f"类型：{note.get('type', 'N/A')}")
            
            create_time = note.get("create_time")
            if create_time:
                dt = datetime.fromtimestamp(create_time)
                output.append(f"发布时间：{dt.strftime('%Y-%m-%d %H:%M:%S')}")
            
            time_desc = note.get("time_desc")
            if time_desc:
                output.append(f"时间描述：{time_desc}")
            
            output.append("-" * 80)
        
        return "\n".join(output)
    
    def export_to_json(self, notes_result: Dict[str, Any], filename: str) -> bool:
        """
        导出笔记数据到 JSON 文件
        
        Args:
            notes_result: 笔记数据结果
            filename: 输出文件名
        
        Returns:
            是否导出成功
        """
        if not notes_result.get("success"):
            return False
        
        try:
            data = notes_result.get("data", {})
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return True
        except Exception:
            return False
    
    def export_to_csv(self, notes_result: Dict[str, Any], filename: str) -> bool:
        """
        导出笔记数据到 CSV 文件
        
        Args:
            notes_result: 笔记数据结果
            filename: 输出文件名
        
        Returns:
            是否导出成功
        """
        if not notes_result.get("success"):
            return False
        
        try:
            import csv
            
            data = notes_result.get("data", {})
            notes = data.get("notes", [])
            
            if not notes:
                return False
            
            # CSV 字段
            fieldnames = [
                'note_id', 'title', 'desc', 'type', 'create_time', 
                'likes', 'collected_count', 'comments_count', 'share_count',
                'user_id', 'nickname', 'images_count'
            ]
            
            with open(filename, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                
                for note in notes:
                    user = note.get("user", {})
                    images = note.get("images_list", [])
                    
                    row = {
                        'note_id': note.get('id', ''),
                        'title': note.get('title', note.get('display_title', '')),
                        'desc': note.get('desc', ''),
                        'type': note.get('type', ''),
                        'create_time': note.get('create_time', ''),
                        'likes': note.get('likes', 0),
                        'collected_count': note.get('collected_count', 0),
                        'comments_count': note.get('comments_count', 0),
                        'share_count': note.get('share_count', 0),
                        'user_id': user.get('userid', ''),
                        'nickname': user.get('nickname', ''),
                        'images_count': len(images)
                    }
                    writer.writerow(row)
            
            return True
        except Exception:
            return False
    
    def export_to_excel(self, notes: List[Dict[str, Any]], filename: str) -> bool:
        """
        导出笔记数据到 Excel 文件
        
        Args:
            notes: 笔记列表
            filename: 输出文件名
        
        Returns:
            是否导出成功
        """
        try:
            import pandas as pd
            
            # 准备数据
            data = []
            for note in notes:
                # 转换时间戳为日期格式
                create_time = note.get('create_time')
                if create_time and isinstance(create_time, (int, float)):
                    try:
                        create_time_str = datetime.fromtimestamp(create_time).strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        create_time_str = str(create_time)
                else:
                    create_time_str = str(create_time) if create_time else ''
                
                row = {
                    '笔记 ID': note.get('id'),
                    '标题': note.get('title', note.get('display_title', '')),
                    '描述': note.get('desc', ''),
                    '类型': note.get('type'),
                    '发布时间': create_time_str,
                    '点赞数': note.get('likes', 0),
                    '收藏数': note.get('collected_count', 0),
                    '评论数': note.get('comments_count', 0),
                    '分享数': note.get('share_count', 0),
                    '作者 ID': note.get('user', {}).get('userid'),
                    '作者昵称': note.get('user', {}).get('nickname'),
                    '图片数量': len(note.get('images_list', [])),
                    '封面图': note.get('images_list', [{}])[0].get('url', '') if note.get('images_list') else ''
                }
                data.append(row)
            
            # 创建 DataFrame
            df = pd.DataFrame(data)
            
            # 导出到 Excel
            df.to_excel(filename, index=False, engine='openpyxl')
            
            # 调整列宽和自动换行
            from openpyxl import load_workbook
            wb = load_workbook(filename)
            ws = wb.active
            
            # 设置所有列为自动换行
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = cell.alignment.copy(wrap_text=True)
            
            # 保存
            wb.save(filename)
            
            return True
        except Exception as e:
            print(f"Excel 导出错误：{e}")
            return False


def main():
    """主函数 - 示例用法"""
    api_token = input("请输入您的 API Token: ").strip()
    
    if not api_token:
        print("❌ API Token 不能为空！")
        return
    
    crawler = XHSUserNotes(api_token)
    
    print("\n" + "=" * 80)
    print("小红书用户笔记采集系统")
    print("=" * 80)
    
    user_id = input("\n请输入用户 ID: ").strip()
    if not user_id:
        print("❌ 用户 ID 不能为空！")
        return
    
    print("\n选择采集模式：")
    print("1. 单页采集（仅第一页）")
    print("2. 全量采集（所有笔记，自动分页）")
    
    mode = input("\n请选择 (1-2, 默认 1): ").strip()
    
    if mode == "2":
        max_pages_input = input("请输入最大页数（留空表示不限制）: ").strip()
        max_pages = int(max_pages_input) if max_pages_input else None
        
        delay_input = input("请输入每页延迟时间（秒，默认 1.0）: ").strip()
        delay = float(delay_input) if delay_input else 1.0
        
        print(f"\n🔍 正在采集用户 {user_id} 的所有笔记...")
        print("请稍候，这可能需要一些时间...\n")
        
        result = crawler.get_all_user_notes(user_id, max_pages, delay)
    else:
        print(f"\n🔍 正在采集用户 {user_id} 的笔记（第一页）...\n")
        result = crawler.get_user_notes(user_id)
    
    print(crawler.format_notes(result))
    
    # 询问是否导出
    if result.get("success"):
        export = input("\n是否导出数据？(y/n): ").strip().lower()
        if export == 'y':
            format_type = input("选择导出格式 (json/csv): ").strip().lower()
            if format_type == 'json':
                filename = f"user_{user_id}_notes.json"
                if crawler.export_to_json(result, filename):
                    print(f"✅ 已导出到：{filename}")
                else:
                    print("❌ 导出失败")
            elif format_type == 'csv':
                filename = f"user_{user_id}_notes.csv"
                if crawler.export_to_csv(result, filename):
                    print(f"✅ 已导出到：{filename}")
                else:
                    print("❌ 导出失败")
            else:
                print("❌ 不支持的格式")


if __name__ == "__main__":
    main()
